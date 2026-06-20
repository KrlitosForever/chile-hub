"""Extrae indicadores urbanos SIEDU desde la Matriz de Indicadores INE.

Fuente: matriz-siedu-publicacion.xlsm (INE, ~505 KB).
  - 5 hojas de medición: LÍNEA DE BASE (2018), SEGUNDA (2019), TERCERA (2020),
    CUARTA (2021), QUINTA (2022).
  - Estructura wide → pivot a long → deduplicación (año más reciente por indicador/comuna).
  - Cobertura parcial esperada: ~117 comunas urbanas, 68 indicadores, ~6700 registros.
"""

import datetime
import os
import sys
from pathlib import Path
from typing import Any

import openpyxl
import polars as pl
import requests

UTC = datetime.timezone.utc

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../.."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

try:
    from src.extractors.base import (
        BaseExtractor,
        ensure_staging_directories,
        write_staging_metadata,
    )
    from src.extractors.source_adapter import build_standard_metadata
except ModuleNotFoundError:
    from base import BaseExtractor, ensure_staging_directories, write_staging_metadata
    from source_adapter import build_standard_metadata

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
RAW_DIR = DATA_DIR / "raw"
STAGING_DIR = DATA_DIR / "staging"
STAGING_CSV_PATH = STAGING_DIR / "indicadores_urbanos_siedu.csv"
METADATA_PATH = STAGING_DIR / "indicadores_urbanos_siedu.metadata.json"

DOWNLOAD_URL = (
    "https://www.ine.gob.cl/docs/default-source/sistema-de-indicadores-y-estandares"
    "-de-desarrollo-urbano/indicadores/actualizaci%C3%B3n-2019/matriz-siedu-publicacion.xlsm"
)
SOURCE_URL = DOWNLOAD_URL
XLSM_FILENAME = "siedu_matriz_indicadores.xlsm"

# Hojas de medición y el año que representa cada una
_SHEET_YEARS: dict[str, int] = {
    "LÍNEA DE BASE": 2018,
    "SEGUNDA MEDICIÓN": 2019,
    "TERCERA MEDICIÓN": 2020,
    "CUARTA MEDICIÓN": 2021,
    "QUINTA MEDICIÓN": 2022,
}

_CATEGORIA_MAP: dict[str, str] = {
    "BPU": "Bienes Públicos Urbanos",
    "DE": "Desplazamientos",
    "EA": "Energía y Ambiente",
    "IS": "Inclusión Social",
    "IP": "Información y Participación",
    "IG": "Integración y Gobernanza",
}

REUSE_POLICY = {
    "status": "open-attribution",
    "license": "Licencia de Datos Abiertos INE",
    "license_url": "https://www.ine.gob.cl/terminos-de-uso",
    "attribution_required": True,
    "redistribution_ok": True,
    "summary": "Indicadores urbanos SIEDU publicados por INE para comunas urbanas seleccionadas.",
}

FALLBACK_ROWS = [
    {
        "anio": 2022,
        "codigo_comuna": "13101",
        "codigo_indicador": "BPU_29",
        "nombre_indicador": "Superficie de áreas verdes públicas por habitante",
        "categoria": "Bienes Públicos Urbanos",
        "valor": 71.4,
        "unidad": "m²/hab",
        "fuente_original": "SIEDU INE",
        "cobertura_tipo": "urbana",
    },
    {
        "anio": 2022,
        "codigo_comuna": "05109",
        "codigo_indicador": "BPU_29",
        "nombre_indicador": "Superficie de áreas verdes públicas por habitante",
        "categoria": "Bienes Públicos Urbanos",
        "valor": 65.8,
        "unidad": "m²/hab",
        "fuente_original": "SIEDU INE",
        "cobertura_tipo": "urbana",
    },
    {
        "anio": 2022,
        "codigo_comuna": "08101",
        "codigo_indicador": "BPU_29",
        "nombre_indicador": "Superficie de áreas verdes públicas por habitante",
        "categoria": "Bienes Públicos Urbanos",
        "valor": 62.3,
        "unidad": "m²/hab",
        "fuente_original": "SIEDU INE",
        "cobertura_tipo": "urbana",
    },
    {
        "anio": 2022,
        "codigo_comuna": "13101",
        "codigo_indicador": "DE_1",
        "nombre_indicador": "Tiempo medio de viaje urbano",
        "categoria": "Desplazamientos",
        "valor": 42.0,
        "unidad": "minutos",
        "fuente_original": "SIEDU INE",
        "cobertura_tipo": "urbana",
    },
    {
        "anio": 2022,
        "codigo_comuna": "05109",
        "codigo_indicador": "DE_1",
        "nombre_indicador": "Tiempo medio de viaje urbano",
        "categoria": "Desplazamientos",
        "valor": 36.0,
        "unidad": "minutos",
        "fuente_original": "SIEDU INE",
        "cobertura_tipo": "urbana",
    },
    {
        "anio": 2022,
        "codigo_comuna": "08101",
        "codigo_indicador": "DE_1",
        "nombre_indicador": "Tiempo medio de viaje urbano",
        "categoria": "Desplazamientos",
        "valor": 34.0,
        "unidad": "minutos",
        "fuente_original": "SIEDU INE",
        "cobertura_tipo": "urbana",
    },
]


def _parse_sheet(ws, year: int) -> list[dict[str, Any]]:
    """Parsea una hoja de medición SIEDU (formato wide) a registros long."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 15:
        return []

    # Row indices (0-based): 10=IDs, 12=units, 13=names, 14+=data
    id_row = rows[10]
    unit_row = rows[12]
    name_row = rows[13]

    indicator_ids = [str(v).strip() for v in id_row[7:] if v is not None]
    n_indicators = len(indicator_ids)
    indicator_names = [
        str(name_row[7 + i]).strip().rstrip(".") if name_row[7 + i] is not None else ""
        for i in range(n_indicators)
    ]
    indicator_units = [
        str(unit_row[7 + i]).strip() if unit_row[7 + i] is not None else ""
        for i in range(n_indicators)
    ]

    records = []
    for row in rows[14:]:
        cut = row[6]
        if cut is None:
            continue
        cut_str = str(cut).strip()
        if not cut_str.isdigit():
            continue
        codigo_comuna = cut_str.zfill(5)

        for i, ind_id in enumerate(indicator_ids):
            val = row[7 + i] if (7 + i) < len(row) else None
            if val is None:
                continue
            val_str = str(val).strip()
            if val_str in ("Sin medición", "Sin valor", ""):
                continue
            try:
                val_float = float(val_str.replace(",", "."))
            except (ValueError, TypeError):
                continue

            prefix = ind_id.split("_")[0]
            categoria = _CATEGORIA_MAP.get(prefix, prefix)
            records.append(
                {
                    "anio": year,
                    "codigo_comuna": codigo_comuna,
                    "codigo_indicador": ind_id,
                    "nombre_indicador": indicator_names[i],
                    "categoria": categoria,
                    "valor": val_float,
                    "unidad": indicator_units[i],
                    "fuente_original": "SIEDU INE",
                    "cobertura_tipo": "urbana",
                }
            )
    return records


def _parse_siedu_xlsm(xlsm_path: Path) -> list[dict[str, Any]]:
    """Parsea el XLSM completo y deduplica conservando el año más reciente."""
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, data_only=True, keep_vba=False)

    all_records: list[dict[str, Any]] = []
    for sheet_name, year in _SHEET_YEARS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        recs = _parse_sheet(ws, year)
        print(f"  {sheet_name} ({year}): {len(recs)} registros")
        all_records.extend(recs)

    # Deduplicar: mantener el año más reciente por (codigo_comuna, codigo_indicador)
    latest: dict[tuple[str, str], dict[str, Any]] = {}
    for rec in all_records:
        key = (rec["codigo_comuna"], rec["codigo_indicador"])
        if key not in latest or rec["anio"] > latest[key]["anio"]:
            latest[key] = rec

    return list(latest.values())


def fetch_data(source_url: str = DOWNLOAD_URL) -> tuple[list[dict[str, Any]], str, str, list[str]]:
    """Descarga la Matriz SIEDU xlsm desde INE y la parsea a nivel comunal."""
    ensure_staging_directories()
    notes: list[str] = [
        "partial_urban_coverage_expected",
        "deduplicado_anno_mas_reciente_por_indicador_comuna",
        "5_mediciones_2018_2022_consolidadas",
    ]

    xlsm_path = RAW_DIR / XLSM_FILENAME

    try:
        print(f"Descargando {source_url} ...")
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
            )
        }
        with requests.get(source_url, headers=headers, timeout=60) as r:
            r.raise_for_status()
            xlsm_path.write_bytes(r.content)
        size_kb = xlsm_path.stat().st_size // 1024
        print(f"Descarga completada ({size_kb} KB).")

        print("Parseando hojas de medición ...")
        rows = _parse_siedu_xlsm(xlsm_path)

        if not rows:
            raise ValueError("El archivo XLSM no produjo registros válidos")

        n_comunas = len({r["codigo_comuna"] for r in rows})
        n_indicadores = len({r["codigo_indicador"] for r in rows})
        notes.append(
            f"live_data: xlsm parseado, {len(rows)} registros, "
            f"{n_comunas} comunas, {n_indicadores} indicadores"
        )
        print(
            f"Extracción completada: {len(rows)} registros ({n_comunas} comunas, {n_indicadores} indicadores)."
        )
        return rows, "live", source_url, notes

    except Exception as exc:
        print(f"Error en extracción live: {exc}. Usando fallback ...")
        notes.append(f"fallback_curated_rows_used: {type(exc).__name__}: {exc}")
        return FALLBACK_ROWS, "fallback", source_url, notes


def normalize_rows(rows: list[dict[str, Any]]) -> pl.DataFrame:
    return (
        pl.DataFrame(rows)
        .with_columns(
            pl.col("anio").cast(pl.Int32),
            pl.col("codigo_comuna").cast(pl.String).str.zfill(5),
            pl.col("codigo_indicador").cast(pl.String),
            pl.col("nombre_indicador").cast(pl.String),
            pl.col("categoria").cast(pl.String),
            pl.col("valor").cast(pl.Float64),
            pl.col("unidad").cast(pl.String),
            pl.col("fuente_original").cast(pl.String),
            pl.col("cobertura_tipo").cast(pl.String),
        )
        .sort(["anio", "codigo_comuna", "codigo_indicador"])
    )


def build_metadata(df: pl.DataFrame, source_mode: str, source_url: str, notes: list[str]) -> dict:
    commune_count = df["codigo_comuna"].n_unique()
    source_detail = (
        "ine_siedu_xlsm_cinco_mediciones_2018_2022"
        if source_mode == "live"
        else "curated_fallback_partial_urban_coverage"
    )
    metadata = build_standard_metadata(
        dataset="indicadores_urbanos_siedu",
        source_name="INE - Sistema de Indicadores y Estándares de Desarrollo Urbano",
        source_url=source_url,
        source_mode=source_mode,
        source_detail=source_detail,
        df=df,
        notes=notes,
        reuse_policy=REUSE_POLICY,
    )
    metadata["coverage"] = {
        "status": "partial_expected",
        "coverage_ratio": round(commune_count / 346, 4),
        "expected_scope": "Comunas urbanas incluidas por SIEDU, no las 346 comunas del país.",
    }
    return metadata


def process_siedu() -> str:
    raw_rows, source_mode, source_url, notes = fetch_data()
    df = normalize_rows(raw_rows)
    metadata = build_metadata(df, source_mode, source_url, notes)
    validation = SieduExtractor().validate(df, metadata)
    if validation["status"] == "error":
        raise SystemExit(f"Validación fallida: {validation['errors']}")
    SieduExtractor().write_staging(df, metadata)
    print(
        f"Indicadores SIEDU guardados en: {STAGING_CSV_PATH} ({df.height} registros, {source_mode})"
    )
    return str(STAGING_CSV_PATH)


class SieduExtractor(BaseExtractor):
    @property
    def dataset_name(self) -> str:
        return "indicadores_urbanos_siedu"

    def fetch(self, **kwargs):
        return fetch_data(**kwargs)

    def normalize(self, raw_data):
        return normalize_rows(raw_data[0])

    def validate(self, df, metadata: dict) -> dict:
        from src.validation import validate_indicadores_urbanos_siedu

        return validate_indicadores_urbanos_siedu(df, metadata)

    def write_staging(self, df, metadata: dict) -> Path:
        ensure_staging_directories()
        df.write_csv(STAGING_CSV_PATH)
        write_staging_metadata(str(METADATA_PATH), metadata)
        return STAGING_CSV_PATH


if __name__ == "__main__":
    process_siedu()
